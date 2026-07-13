from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from speccheck.main import summary
from speccheck.report import get_default_template_path


def _build_speccheck_summary_input(source_csv, destination):
    source = pd.read_csv(source_csv)
    records = []
    for _, row in source.iterrows():
        qc_pass = str(row["qc_verdict"]).lower() == "pass"
        n50_pass = str(row["N50_verdict"]).lower() == "pass"
        contig_pass = str(row["no_of_contigs_verdict"]).lower() == "pass"
        genome_pass = str(row["Genome_Size_verdict"]).lower() == "pass"
        checkm_pass = qc_pass and float(row["Contamination"]) <= 2.0
        records.append(
            {
                "sample_id": row["sample"],
                "all_checks_passed": qc_pass,
                "Checkm.all_checks_passed": checkm_pass,
                "Checkm.Completeness": row["Completeness_Specific"],
                "Checkm.Contamination": row["Contamination"],
                "Checkm.GC": row["GC_Content"],
                "Checkm.Genome size (bp)": row["Genome_Size"],
                "Checkm.# contigs": row["number"],
                "Checkm.N50 (scaffolds)": row["N50"],
                "Checkm.Total_Coding_Sequences": row["Total_Coding_Sequences"],
                "Checkm.Completeness.check": True,
                "Checkm.Contamination.check": float(row["Contamination"]) <= 2.0,
                "Quast.all_checks_passed": n50_pass and contig_pass and genome_pass,
                "Quast.N50": row["N50"],
                "Quast.N50.check": n50_pass,
                "Quast.# contigs (>= 0 bp)": row["number"],
                "Quast.# contigs (>= 0 bp).check": contig_pass,
                "Quast.Total length (>= 0 bp)": row["Genome_Size"],
                "Quast.Total length": row["Genome_Size"],
                "Quast.Total length (>= 0 bp).check": genome_pass,
                "Quast.GC (%)": row["GC_Content"],
                "Quast.GC (%).check": str(row["GC_Content_verdict"]).lower() == "pass",
                "Quast.Largest contig": row["longest"],
                "Speciator.all_checks_passed": True,
                "Speciator.speciesName": row["species_sylph"],
                "Speciator.confidence": "good",
                "Sylph.all_checks_passed": True,
                "Sylph.top_species": row["species_sylph"],
                "Sylph.top_taxonomic_abundance": 0.95 if qc_pass else 0.75,
                "Sylph.number_of_genomes": 1,
            }
        )
    pd.DataFrame(records).to_csv(destination, index=False)


def test_summary_generates_interactive_html_and_xlsx(tmp_path):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()

    _build_speccheck_summary_input("tests/qualibact/ecoli_pass_subset.csv", input_dir / "pass.csv")
    _build_speccheck_summary_input("tests/qualibact/ecoli_fail_subset.csv", input_dir / "fail.csv")

    xlsx_output = output_dir / "report.xlsx"
    summary(
        str(input_dir),
        str(output_dir),
        "Speciator.speciesName",
        "sample_id",
        get_default_template_path(),
        plot=True,
        xlsx_output=str(xlsx_output),
        qualifyr_style=True,
    )

    report_html = (output_dir / "report.html").read_text(encoding="utf-8")
    assert (output_dir / "report.csv").exists()
    assert (output_dir / "report.html").exists()
    assert not (output_dir / "bulma.css").exists()
    assert xlsx_output.exists()
    assert "table-filter" in report_html
    assert "qualifyr-like layout" in report_html
    assert "Speciator" in report_html
    assert "Confidence" in report_html
    assert '<link rel="stylesheet" href="bulma.css">' not in report_html
    assert ".report-header" in report_html

    workbook = openpyxl.load_workbook(xlsx_output)
    assert "report" in workbook.sheetnames
    assert "qc_status" in workbook.sheetnames


def test_summary_respects_no_interactive_tables(tmp_path):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()

    _build_speccheck_summary_input("tests/qualibact/ecoli_pass_subset.csv", input_dir / "pass.csv")

    summary(
        str(input_dir),
        str(output_dir),
        "Speciator.speciesName",
        "sample_id",
        get_default_template_path(),
        plot=True,
        interactive_tables=False,
        qualifyr_style=True,
    )

    report_html = (output_dir / "report.html").read_text(encoding="utf-8")
    assert 'class="table-filter"' not in report_html
    assert 'class="table report-table js-sort-filter"' not in report_html
    assert "parseValue" not in report_html
    assert '<link rel="stylesheet" href="bulma.css">' not in report_html


def test_summary_adds_qualibact_compatibility_columns(tmp_path):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    pd.DataFrame(
        [
            {
                "sample_id": "PASS1",
                "all_checks_passed": True,
                "Checkm.all_checks_passed": True,
                "Checkm.Completeness": 100,
                "Checkm.Contamination": 0.2,
                "Checkm.GC_Content": 50.5,
                "Checkm.Genome_Size": 5100000,
                "Checkm.Contig_N50": 120000,
                "Checkm.Total_Contigs": 200,
                "Quast.all_checks_passed": True,
                "Quast.N50": 120000,
                "Quast.# contigs (>= 0 bp)": 200,
                "Quast.GC (%)": 50.5,
                "Quast.Total length (>= 0 bp)": 5100000,
                "Quast.Largest contig": 500000,
                "Checkm.Total_Coding_Sequences": 4800,
                "Speciator.speciesName": "Escherichia coli",
                "Speciator.confidence": "good",
            },
            {
                "sample_id": "WARN1",
                "all_checks_passed": True,
                "Checkm.all_checks_passed": True,
                "Checkm.Completeness": 100,
                "Checkm.Contamination": 0.3,
                "Checkm.GC_Content": 50.4387,
                "Checkm.Genome_Size": 5695351,
                "Checkm.Contig_N50": 114262,
                "Checkm.Total_Contigs": 665,
                "Quast.all_checks_passed": True,
                "Quast.N50": 114262,
                "Quast.# contigs (>= 0 bp)": 665,
                "Quast.GC (%)": 50.4387,
                "Quast.Total length (>= 0 bp)": 5695351,
                "Quast.Largest contig": 297694,
                "Checkm.Total_Coding_Sequences": 6055,
                "Speciator.speciesName": "Escherichia coli",
                "Speciator.confidence": "good",
            },
            {
                "sample_id": "FAIL1",
                "all_checks_passed": True,
                "Checkm.all_checks_passed": True,
                "Checkm.Completeness": 100,
                "Checkm.Contamination": 0.4,
                "Checkm.GC_Content": 50.5585,
                "Checkm.Genome_Size": 5738641,
                "Checkm.Contig_N50": 126621,
                "Checkm.Total_Contigs": 679,
                "Quast.all_checks_passed": True,
                "Quast.N50": 126621,
                "Quast.# contigs (>= 0 bp)": 679,
                "Quast.GC (%)": 50.5585,
                "Quast.Total length (>= 0 bp)": 5738641,
                "Quast.Largest contig": 312707,
                "Checkm.Total_Coding_Sequences": 6105,
                "Speciator.speciesName": "Escherichia coli",
                "Speciator.confidence": "good",
            },
        ]
    ).to_csv(input_dir / "samples.csv", index=False)

    summary(
        str(input_dir),
        str(output_dir),
        "Speciator.speciesName",
        "sample_id",
        get_default_template_path(),
        plot=True,
        qualifyr_style=True,
        qualibact_compat=True,
    )

    report = pd.read_csv(output_dir / "report.csv")
    tiers = dict(zip(report["sample_id"], report["qualibact_compat_tier"], strict=False))
    passed = dict(zip(report["sample_id"], report["all_checks_passed"], strict=False))
    html = (output_dir / "report.html").read_text(encoding="utf-8")

    assert tiers == {"PASS1": "PASS", "WARN1": "WARN", "FAIL1": "FAIL"}
    assert passed["WARN1"] == "PASSED"
    assert passed["FAIL1"] == "FAILED"
    assert "qualibact_compat_tier" in html
    assert "Total_Coding_Sequences &gt;5800.0" in html


def test_summary_rejects_duplicate_sample_ids(tmp_path):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    pd.DataFrame(
        [
            {"sample_id": "S1", "all_checks_passed": True},
            {"sample_id": "S1", "all_checks_passed": False},
        ]
    ).to_csv(input_dir / "duplicates.csv", index=False)

    with pytest.raises(ValueError, match="duplicate sample IDs"):
        summary(
            str(input_dir),
            str(output_dir),
            "Speciator.speciesName",
            "sample_id",
            get_default_template_path(),
            plot=False,
        )


def test_summary_rejects_missing_sample_column(tmp_path):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    pd.DataFrame([{"wrong_id": "S1", "all_checks_passed": True}]).to_csv(
        input_dir / "missing_sample.csv", index=False
    )

    with pytest.raises(ValueError, match="missing required sample column"):
        summary(
            str(input_dir),
            str(output_dir),
            "Speciator.speciesName",
            "sample_id",
            get_default_template_path(),
            plot=False,
        )


def test_summary_ignores_detailed_csv_and_output_directory(tmp_path):
    input_dir = tmp_path / "input"
    output_dir = input_dir / "summary"
    input_dir.mkdir()
    pd.DataFrame([{"sample_id": "S1", "all_checks_passed": True}]).to_csv(
        input_dir / "sample.csv", index=False
    )
    pd.DataFrame([{"sample_id": "S1", "extra": "legacy"}]).to_csv(
        input_dir / "detailed.sample.csv", index=False
    )
    output_dir.mkdir()
    pd.DataFrame([{"sample_id": "S1", "old": "report"}]).to_csv(
        output_dir / "report.csv", index=False
    )

    summary(
        str(input_dir),
        str(output_dir),
        "Speciator.speciesName",
        "sample_id",
        get_default_template_path(),
        plot=False,
    )

    report = pd.read_csv(output_dir / "report.csv")
    assert list(report["sample_id"]) == ["S1"]
    assert "extra" not in report.columns
    assert "old" not in report.columns
